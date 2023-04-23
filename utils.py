from openpyxl import load_workbook
import sqlite3
import os

USER_DIR = os.path.expanduser("~\AppData\Local")
DATA_DIR = os.path.join(USER_DIR, "TinFetcher")
DATABASE_DIR = os.path.join(DATA_DIR, "database")

def database_file():
    if not os.path.exists(DATABASE_DIR):
        os.makedirs(DATABASE_DIR)

    database = os.path.join(DATABASE_DIR, "client.db")

    return database

def set_database_dir(path):
    global DATABASE_DIR
    
    if not os.path.isdir(path):
        return
    
    DATABASE_DIR = path

def get_database_dir():
    return DATABASE_DIR

def set_database_file(file):
    global database
    
    if not os.path.isfile(file):
        return
    
    database = file
    set_database_dir(os.path.dirname(database))

def get_database_file():
    return database

def create_connection(db_file):
    connection = None

    try:
        connection = sqlite3.connect(db_file)
        cursor = connection.cursor()
        cursor.execute(""" CREATE TABLE IF NOT EXISTS clients (tin TEXT NOT NULL UNIQUE PRIMARY KEY, name TEXT NOT NULL) """)
        connection.commit()
        cursor.close()

    except sqlite3.Error as e:
        print(f"The Error '{e}' occurred.")

    return connection

def fetch_clients(conn, client_name=None) -> list:
    cur = conn.cursor()

    if not client_name:
        cur.execute(""" SELECT * FROM clients """)
    else:
        cur.execute(""" SELECT * FROM clients WHERE name=? """, (client_name, ))
    
    data = cur.fetchall()
    
    return data

def initialize(conn):
    cur = conn.cursor()

    cur.execute(""" CREATE TABLE IF NOT EXISTS clients ( tin TEXT NOT NULL UNIQUE, name TEXT NOT NULL ) """)
    excel_file = os.path.join(eBir_path, "Clients.xlsx")

    wb = load_workbook(excel_file)
    ws = wb.active
    tin = [row.value for row in ws["A"][1:]]
    clients = [row.value for row in ws["B"][1:]]

    for tin, clients in zip(tin, clients):
        try:
            cur.execute(""" INSERT INTO clients (tin, name) VALUES (?, ?) """, (tin, clients))
        except:
            continue

    conn.commit()
    cur.close()


if __name__ == "__main__":
    eBir_path = "E:\eBIRForms"

    pathExists = os.path.exists(eBir_path)

    if not pathExists:
        raise Exception("Path does not exist: %s" % eBir_path)

    DB_FILE = database_file()

    with create_connection(DB_FILE) as conn:
        initialize(conn)